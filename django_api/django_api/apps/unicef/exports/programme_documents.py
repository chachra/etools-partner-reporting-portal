import hashlib
import logging
import os
import tempfile

from django.http import HttpResponse
from django.utils import timezone
from easy_pdf.exceptions import PDFRenderingError
from easy_pdf.rendering import render_to_pdf, make_response
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD, FORMAT_PERCENTAGE
from openpyxl.utils import get_column_letter

from unicef.exports.utilities import PARTNER_PORTAL_DATE_FORMAT_EXCEL, HTMLTableCell
from unicef.templatetags.pdf_extras import format_currency

logger = logging.getLogger(__name__)


class ProgrammeDocumentsXLSXExporter:

    def __init__(self, programme_documents):
        self.programme_documents = programme_documents
        filename = hashlib.sha256(';'.join([str(p.pk) for p in programme_documents]).encode('utf-8')).hexdigest()
        self.file_path = os.path.join(tempfile.gettempdir(), filename + '.xlsx')
        self.workbook = Workbook()
        self.worksheet = self.workbook.get_active_sheet()
        self.display_name = '[{:%a %-d %b %-H-%M-%S %Y}] {} Programme Document(s) Summary.xlsx'.format(
            timezone.now(), programme_documents.count()
        )

        self.header_style = NamedStyle(name="header")
        self.header_style.font = Font(bold=True, color='FFFFFF')
        self.header_style.fill = PatternFill("solid", fgColor="3195EE")
        self.header_style.alignment = Alignment(horizontal='center', vertical='justify')

    def fill_worksheet(self):
        self.worksheet.title = 'Programme Document(s) Summary'
        headers = [
            'Title',
            'Agreement',
            'Document Type',
            'Reference Number',
            'UNICEF Office(s)',
            'UNICEF Focal Point(s)',
            'Partner Focal Point(s)',
            'In response to an HRP',

            'PD/SSFA status',
            'Start date',
            'End date',
            'CSO contribution',
            'Total UNICEF cash',
            'Total UNICEF supplies',
            'Total Budget',
            'Cash Transfers to Date ($)',
            'Cash Transfers to Date (%)',
        ]

        current_row = 1
        column_widths = []

        for column, header_text in enumerate(headers):
            column_widths.append(len(header_text))
            column += 1  # columns are not 0-indexed...
            cell = self.worksheet.cell(row=current_row, column=column, value=header_text)
            cell.style = self.header_style
        current_row += 1

        for pd in self.programme_documents:
            if pd.budget:
                funds_received_to_date_percentage = pd.funds_received_to_date / pd.budget
            else:
                funds_received_to_date_percentage = 0

            data_row = [
                (pd.title, None),
                (pd.agreement, None),
                (pd.get_document_type_display(), None),
                (pd.reference_number, None),
                (pd.unicef_office, None),
                (', '.join([person.name for person in pd.unicef_focal_point.all()]), None),
                (', '.join([person.name for person in pd.partner_focal_point.all()]), None),
                (None, None),  # This field is not calculated anywhere yet
                (pd.get_status_display(), None),
                (pd.start_date, PARTNER_PORTAL_DATE_FORMAT_EXCEL),
                (pd.end_date, PARTNER_PORTAL_DATE_FORMAT_EXCEL),
                (pd.cso_contribution, FORMAT_CURRENCY_USD),
                (pd.total_unicef_cash, FORMAT_CURRENCY_USD),
                (pd.in_kind_amount, FORMAT_CURRENCY_USD),
                (pd.budget, FORMAT_CURRENCY_USD),
                (pd.funds_received_to_date, FORMAT_CURRENCY_USD),
                (funds_received_to_date_percentage, FORMAT_PERCENTAGE),
            ]

            if not len(headers) == len(data_row):
                raise Exception('Header and data row length mismatch!')

            for column, (cell_data, cell_format) in enumerate(data_row):
                try:
                    column_widths[column] = max(column_widths[column], len(cell_data) + 2)
                except TypeError:
                    column_widths[column] = max(column_widths[column], len(str(cell_data)) + 2)
                column += 1  # columns are not 0-indexed...
                cell = self.worksheet.cell(row=current_row, column=column, value=cell_data)
                if cell_format:
                    cell.number_format = cell_format

            current_row += 1

        for column, width in enumerate(column_widths):
            column += 1
            self.worksheet.column_dimensions[get_column_letter(column)].width = width

        self.workbook.save(self.file_path)

    def get_as_response(self):
        self.fill_worksheet()
        response = HttpResponse()
        response.content_type = self.worksheet.mime_type
        with open(self.file_path, 'rb') as content:
            response.write(content.read())
        self.cleanup()
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(self.display_name)
        return response

    def cleanup(self):
        os.remove(self.file_path)


class ProgrammeDocumentsPDFExporter:

    template_name = 'programme_documents_pdf_export.html'

    def __init__(self, programme_documents):
        self.programme_documents = programme_documents
        self.file_name = '[{:%a %-d %b %-H-%M-%S %Y}] {} Programme Document(s) Summary.pdf'.format(
            timezone.now(), programme_documents.count()
        )

    def get_context(self):
        context = {
            'title': 'Programme Document(s) Summary',
        }

        rows = [[
            HTMLTableCell('Title', element='th'),
            HTMLTableCell('General Info', colspan=4, element='th'),
            HTMLTableCell('Financial Info', colspan=2, element='th'),
        ]]

        for pd in self.programme_documents.order_by('id'):
            if pd.budget:
                funds_received_to_date_percentage = int(round(pd.funds_received_to_date * 100 / pd.budget, 0))
            else:
                funds_received_to_date_percentage = 0

            rows.append([
                HTMLTableCell(pd.title, rowspan=5),
                HTMLTableCell('Agreement', element='th'),
                HTMLTableCell(pd.agreement),
                HTMLTableCell('UNICEF Office(s)', element='th'),
                HTMLTableCell(pd.unicef_office),
                HTMLTableCell('CSO contribution', element='th'),
                HTMLTableCell(format_currency(pd.cso_contribution, pd.cso_contribution_currency)),
            ])

            rows.append([
                HTMLTableCell('Document Type', element='th'),
                HTMLTableCell(pd.get_document_type_display()),
                HTMLTableCell('UNICEF Focal Point(s)', element='th'),
                HTMLTableCell(', '.join([person.name for person in pd.unicef_focal_point.all()])),
                HTMLTableCell('Total UNICEF cash', element='th'),
                HTMLTableCell(format_currency(pd.total_unicef_cash, pd.total_unicef_cash_currency))
            ])
            rows.append([
                HTMLTableCell('Reference Number', element='th'),
                HTMLTableCell(pd.reference_number),
                HTMLTableCell('Partner Focal Point(s)', element='th'),
                HTMLTableCell(', '.join([person.name for person in pd.partner_focal_point.all()])),
                HTMLTableCell('Total UNICEF supplies', element='th'),
                HTMLTableCell(format_currency(pd.in_kind_amount, pd.in_kind_amount_currency))
            ])
            rows.append([
                HTMLTableCell('PD/SSFA status', element='th'),
                HTMLTableCell(pd.get_status_display()),
                HTMLTableCell('Start Date', element='th'),
                HTMLTableCell(pd.start_date),
                HTMLTableCell('Total Budget', element='th'),
                HTMLTableCell(format_currency(pd.budget, pd.budget_currency))
            ])
            rows.append([
                HTMLTableCell('In response to an HRP', element='th'),
                HTMLTableCell(''),
                HTMLTableCell('End Date', element='th'),
                HTMLTableCell(pd.end_date),
                HTMLTableCell('Cash Transfers to Date', element='th'),
                HTMLTableCell('{} ({}%)'.format(
                    format_currency(pd.funds_received_to_date, pd.funds_received_to_date_currency),
                    funds_received_to_date_percentage
                ))
            ])

        context['rows'] = rows
        context['programme_documents'] = self.programme_documents

        return context

    def get_as_response(self):
        try:
            pdf = render_to_pdf(self.template_name, self.get_context())
            response = make_response(pdf)
            response['Content-disposition'] = 'inline; filename="{}"'.format(self.file_name)
            return response
        except PDFRenderingError:
            logger.exception('Error trying to render PDF')
            return HttpResponse('Error trying to render PDF')
